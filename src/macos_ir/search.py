"""Search index over a generated artifacts.xlsx workbook.

Builds an in-memory inverted index on first query and reuses it across
subsequent queries in the same session. Each sheet is treated as one
artifact; results are grouped by sheet with guide blurbs rendered
alongside by the caller.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore

# Split on common separators so "192.168.50.5" and "user@host"
# break into both the whole string and its parts for looser matching.
_TOKEN_SPLIT = re.compile(r"[\s,;|\t\n\r]+")


@dataclass
class SearchHit:
    sheet: str
    row: int
    col: int
    cell_value: str
    preview: str  # truncated row preview for UI


@dataclass
class SearchResults:
    query: str
    hits_by_sheet: dict[str, list[SearchHit]] = field(default_factory=dict)

    @property
    def total_hits(self) -> int:
        return sum(len(h) for h in self.hits_by_sheet.values())

    @property
    def sheet_count(self) -> int:
        return len(self.hits_by_sheet)


class WorkbookSearchIndex:
    """Load a .xlsx once and answer many queries against it."""

    def __init__(self, xlsx_path: Path):
        self.xlsx_path = Path(xlsx_path)
        self._rows: dict[str, list[list[str]]] = {}  # sheet -> rows of cells as strings
        self._headers: dict[str, list[str]] = {}
        self._built = False

    def _ensure_built(self, progress_cb=None) -> None:
        if self._built:
            return
        if load_workbook is None:
            raise RuntimeError("openpyxl not installed")

        wb = load_workbook(self.xlsx_path, read_only=True, data_only=True)
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            headers: list[str] = []
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                row_cells = [("" if v is None else str(v)) for v in row]
                if row_idx == 0:
                    headers = row_cells
                    continue
                rows.append(row_cells)
                if progress_cb and len(rows) % 500 == 0:
                    progress_cb(sheet_name, row_idx)
            self._rows[sheet_name] = rows
            self._headers[sheet_name] = headers
            if progress_cb:
                progress_cb(sheet_name, len(rows))
        wb.close()
        self._built = True

    def sheet_names(self) -> list[str]:
        self._ensure_built()
        return list(self._rows.keys())

    def sheet_row_count(self, sheet: str) -> int:
        self._ensure_built()
        return len(self._rows.get(sheet, []))

    def query(self, term: str, scope_sheet: str | None = None,
              max_per_sheet: int = 100, progress_cb=None) -> SearchResults:
        """Return SearchResults for a case-insensitive substring match of term."""
        self._ensure_built(progress_cb=progress_cb)
        term_lc = term.lower().strip()
        results = SearchResults(query=term)
        if not term_lc:
            return results

        sheets = [scope_sheet] if scope_sheet else self._rows.keys()
        for sheet in sheets:
            rows = self._rows.get(sheet, [])
            hits: list[SearchHit] = []
            for row_idx, row in enumerate(rows):
                row_hit = False
                for col_idx, cell in enumerate(row):
                    if term_lc in cell.lower():
                        preview = self._format_preview(sheet, row)
                        hits.append(SearchHit(
                            sheet=sheet,
                            row=row_idx + 2,          # 1-based + header row
                            col=col_idx + 1,
                            cell_value=cell,
                            preview=preview,
                        ))
                        row_hit = True
                        break  # one hit per row — enough for UI
                if row_hit and len(hits) >= max_per_sheet:
                    break
            if hits:
                results.hits_by_sheet[sheet] = hits
        return results

    def _format_preview(self, sheet: str, row: list[str], max_len: int = 220) -> str:
        """Single-line summary of a row for display — joins non-empty cells."""
        parts = [c for c in row if c]
        joined = " · ".join(parts)
        if len(joined) > max_len:
            joined = joined[:max_len - 1] + "…"
        return joined
