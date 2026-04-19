"""Plugin runner and XLSX workbook writer.

Invokes dissect plugin functions against a target and writes a multi-sheet
XLSX workbook with a summary sheet plus one sheet per artifact.
"""

from __future__ import annotations

import json
import re
import subprocess
import sys
import time
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from collections.abc import Callable

# ──────────────────────────────────────────────────────────────────────────────
# Target resolution
# ──────────────────────────────────────────────────────────────────────────────

DISSECT_PSEUDO_TARGETS = {"local"}


def resolve_target(collection_dir: str) -> str:
    """Resolve the target path (uploads/auto/) from the collection dir."""
    if collection_dir in DISSECT_PSEUDO_TARGETS:
        return collection_dir
    p = Path(collection_dir)
    if (p / "uploads" / "auto").is_dir():
        return str(p / "uploads" / "auto")
    if p.name == "auto" and (p / "Users").is_dir():
        return str(p)
    if (p / "auto").is_dir():
        return str(p / "auto")
    return str(p)


# ──────────────────────────────────────────────────────────────────────────────
# Plugin runner
# ──────────────────────────────────────────────────────────────────────────────

def run_function(
    func: str,
    target: str,
    plugin_path: str,
    timeout: int = 600,
    progress_cb: Callable[[int], None] | None = None,
) -> tuple[list[dict] | None, str | None]:
    """Invoke dissect for one function, streaming records line-by-line.

    Returns (records, err). Uses an IDLE timeout rather than wall-clock
    so long-running plugins keep going as long as they produce output.
    """
    cmd = [
        sys.executable, "-m", "dissect.target.tools.query",
        "--plugin-path", plugin_path,
        "-f", func,
        target, "-j",
    ]

    try:
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
        )
    except OSError as e:
        return [], f"spawn failed: {e}"

    records: list[dict] = []
    last_activity = time.monotonic()
    timed_out = False

    import selectors
    sel = selectors.DefaultSelector()
    sel.register(proc.stdout, selectors.EVENT_READ)

    try:
        while True:
            events = sel.select(timeout=1.0)
            if events:
                line = proc.stdout.readline()
                if not line:
                    break
                last_activity = time.monotonic()
                line = line.strip()
                if not line:
                    continue
                try:
                    obj = json.loads(line)
                except json.JSONDecodeError:
                    continue
                if not isinstance(obj, dict) or obj.get("_type") == "recorddescriptor":
                    continue
                records.append(obj)
                if progress_cb and len(records) % 500 == 0:
                    progress_cb(len(records))
            else:
                if time.monotonic() - last_activity > timeout:
                    timed_out = True
                    break
                if proc.poll() is not None:
                    tail = proc.stdout.read()
                    for tl in (tail.splitlines() if tail else ()):
                        tl = tl.strip()
                        if not tl:
                            continue
                        try:
                            obj = json.loads(tl)
                        except json.JSONDecodeError:
                            continue
                        if not isinstance(obj, dict) or obj.get("_type") == "recorddescriptor":
                            continue
                        records.append(obj)
                    break
    finally:
        sel.close()

    if timed_out:
        try:
            proc.kill()
            proc.wait(timeout=5)
        except Exception:
            pass
        return None, "timeout"

    proc.wait()
    if proc.returncode != 0:
        err = (proc.stderr.read() or "").strip() if proc.stderr else ""
        if "Unsupported plugin" in err or "has no function" in err:
            return [], "incompat"
        if not records:
            return [], err.split("\n")[0][:200] if err else f"exit {proc.returncode}"
    return records, None


# ──────────────────────────────────────────────────────────────────────────────
# XLSX output
# ──────────────────────────────────────────────────────────────────────────────

DISSECT_INTERNAL_FIELDS = {
    "_type", "_classification", "_generated", "_version",
    "_recorddescriptor", "_source", "_classname", "_desc",
}

_INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")
_XLSX_ILLEGAL_RE = re.compile(r"[\000-\010\013\014\016-\037]")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = _INVALID_SHEET_CHARS.sub("_", name)[:31]
    base = cleaned or "sheet"
    candidate = base
    n = 2
    while candidate in used:
        suffix = f"~{n}"
        candidate = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


def _record_columns(records: list[dict]) -> list[str]:
    seen: list[str] = []
    for rec in records:
        for k in rec:
            if k not in DISSECT_INTERNAL_FIELDS and k not in seen:
                seen.append(k)
    return seen


def _scrub(s: str) -> str:
    s = _XLSX_ILLEGAL_RE.sub("", s)
    if len(s) > 32760:
        s = s[:32757] + "..."
    return s


def _coerce_cell(v):
    if v is None:
        return ""
    if isinstance(v, bool) or isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        return _scrub(v)
    if isinstance(v, (list, tuple)):
        return _scrub(", ".join(str(x) for x in v))
    if isinstance(v, dict):
        return _scrub(json.dumps(v, default=str))
    return _scrub(str(v))


def write_xlsx(
    per_source: list[tuple],
    all_records: list[tuple[str, list[dict]]],
    path: str,
) -> None:
    """Write a multi-sheet workbook: Summary + one sheet per function."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_align = Alignment(horizontal="left", vertical="top", wrap_text=False)

    def _write_sheet(ws, columns, rows_iter):
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.freeze_panes = "A2"
        for r in rows_iter:
            ws.append([_coerce_cell(r.get(c)) for c in columns])
        ws.auto_filter.ref = ws.dimensions

    used: set[str] = set()

    # Summary sheet
    ws = wb.create_sheet(_safe_sheet_name("Summary", used))
    summary_cols = ["function", "category", "records", "status"]
    summary_rows = [
        {"function": p[0], "category": p[1], "records": p[2], "status": p[3] or "ok"}
        for p in per_source
    ]
    _write_sheet(ws, summary_cols, summary_rows)

    # Per-artifact sheets
    for func, records in all_records:
        if not records:
            continue
        ws = wb.create_sheet(_safe_sheet_name(func, used))
        cols = _record_columns(records)
        _write_sheet(ws, cols, records)

    wb.save(path)
